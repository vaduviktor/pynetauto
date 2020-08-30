from nornir.core.deserializer.inventory import Inventory
from openpyxl import load_workbook
from openpyxl.utils.exceptions import SheetTitleException


def get_devices_excel() -> dict:
    """
    Get device list from excel file and transform it to Nornir inventory host dict

    """
    dev_dict = {}

    try:
        workbook = load_workbook(filename="SwitchList.xlsx")
        sheet = workbook["Group 2"]
    except FileNotFoundError:
        raise FileNotFoundError("No SwitchList.xlsx excel file found in working dir")
    except (SheetTitleException, KeyError):
        raise SheetTitleException("Sheet not found")

    ipaddr = ''
    for i in range(2, sheet.max_row + 1):
        hostname = sheet[f'A{i}'].value
        # if ipaddr same as previous in the list, its part of the stack, if not add to dict
        if not str(sheet[f'F{i}'].value) == ipaddr:
            ipaddr = str(sheet[f'F{i}'].value)
            dev_dict.update({
                f'{hostname}': {
                    "hostname": ipaddr,
                    "groups": ["ciscoSW"],
                    "data": {
                        "customer": "LAB"
                    }
                }
            })

    return dev_dict


class ExcelInventory(Inventory):
    def __init__(self, **kwargs):
        hosts = get_devices_excel()
        groups = {
            "ciscoSW": {
                "username": '',
                "password": '',
                "platform": 'ios',
                "connection_options": {
                    "netmiko": {
                        "extras": {
                            "timeout": 5,
                            "global_delay_factor": 3,
                            "secret": ''
                        }
                    },
                    "napalm": {
                        "extras": {
                            "timeout": 5,
                            "optional_args": {
                                "secret": '',
                                "global_delay_factor": 3
                            }
                        }
                    }
                },
                "data": {
                    "type": "network_device"
                }
            }
        }
        defaults = {"data": {"location": "LAB", "language": "py"}}

        # passing the data to the parent class so the data is
        # transformed into actual Host/Group objects
        # and set default data for all hosts
        super().__init__(hosts=hosts, groups=groups, defaults=defaults, **kwargs)

from typing import Any, Dict, Type
from nornir.core.inventory import (
    Inventory,
    Group,
    Groups,
    Host,
    Hosts,
    Defaults,
    ConnectionOptions,
    HostOrGroup,
    ParentGroups,
)
from openpyxl import load_workbook
from openpyxl.utils.exceptions import SheetTitleException


def _get_connection_options(data: Dict[str, Any]) -> Dict[str, ConnectionOptions]:
    cp = {}
    for cn, c in data.items():
        cp[cn] = ConnectionOptions(
            hostname=c.get("hostname"),
            port=c.get("port"),
            username=c.get("username"),
            password=c.get("password"),
            platform=c.get("platform"),
            extras=c.get("extras"),
        )
    return cp


def _get_defaults(data: Dict[str, Any]) -> Defaults:
    return Defaults(
        hostname=data.get("hostname"),
        port=data.get("port"),
        username=data.get("username"),
        password=data.get("password"),
        platform=data.get("platform"),
        data=data.get("data"),
        connection_options=_get_connection_options(data.get("connection_options", {})),
    )


def _get_inventory_element(
    typ: Type[HostOrGroup], data: Dict[str, Any], name: str, defaults: Defaults
) -> HostOrGroup:
    return typ(
        name=name,
        hostname=data.get("hostname"),
        port=data.get("port"),
        username=data.get("username"),
        password=data.get("password"),
        platform=data.get("platform"),
        data=data.get("data"),
        groups=data.get(
            "groups"
        ),  # this is a hack, we will convert it later to the correct type
        defaults=defaults,
        connection_options=_get_connection_options(data.get("connection_options", {})),
    )


def _get_devices_excel(filename: str) -> dict:
    """
    Get device list from excel file and transform it to Nornir inventory host dict

    """
    dev_dict = {}

    try:
        workbook = load_workbook(filename=filename)
        sheet = workbook["Group 2"]  # data is in tab Group 2
    except FileNotFoundError:
        raise FileNotFoundError("No SwitchList.xlsx excel file found in working dir")
    except (SheetTitleException, KeyError):
        raise SheetTitleException("Sheet not found")

    ipaddr = ''
    for i in range(2, sheet.max_row + 1):
        hostname = sheet[f'A{i}'].value
        # if ipaddr same as previous in the list, its part of the stack, if not add to dict
        if not str(sheet[f'F{i}'].value) == ipaddr:
            ipaddr = str(sheet[f'F{i}'].value)
            dev_dict.update({
                f'{hostname}': {
                    "hostname": ipaddr,
                    "groups": ["ciscoSW"],
                    "data": {
                        "customer": "LAB"
                    }
                }
            })

    return dev_dict


class ExcelInventory:
    defaults_dict = {"data": {"location": "LAB", "language": "py"}}
    groups_dict = {
        "ciscoSW": {
            "username": 'username',
            "password": 'pass',
            "platform": 'ios',
            "connection_options": {
                "netmiko": {
                    "extras": {
                        "timeout": 5,
                        "global_delay_factor": 3,
                        "secret": ''
                    }
                },
                "napalm": {
                    "extras": {
                        "timeout": 5,
                        "optional_args": {
                            "secret": '',
                            "global_delay_factor": 3
                        }
                    }
                }
            },
            "data": {
                "type": "network_device"
            }
        }
    }

    def __init__(self, excel_filename: str = "SwitchList.xlsx", **kwargs: Any) -> None:
        self._excel_file = excel_filename

    def load(self) -> Inventory:
        defaults = _get_defaults(self.defaults_dict)

        hosts = Hosts()
        hosts_dict = _get_devices_excel(self._excel_file)
        for n, h in hosts_dict.items():
            hosts[n] = _get_inventory_element(Host, h, n, defaults)

        groups = Groups()
        for n, g in self.groups_dict.items():
            groups[n] = _get_inventory_element(Group, g, n, defaults)

        for h in hosts.values():
            h.groups = ParentGroups([groups[g] for g in h.groups])

        for g in groups.values():
            g.groups = ParentGroups([groups[g] for g in g.groups])

        return Inventory(hosts=hosts, groups=groups, defaults=defaults)


